import json

from flask import Flask, render_template, request
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
# from xvfbwrapper import Xvfb
import openpyxl
import time
import logging


logging.basicConfig(filename="log.log",
                    level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

app = Flask(__name__)
LOGGER = logging.getLogger("logger")


@app.route("/", methods=["POST", "GET"])
def index():
    args = {"method": "GET", "file_type": "clear", "rows": "1"}
    if request.method == "POST":
        file = request.files["file"]
        args["method"] = "POST"
        file_type = file.filename.split('.')[-1]
        args['file_type'] = file_type
        if file_type == 'xlsx':
            file.save('TEMP.xlsx')
            wb = openpyxl.load_workbook('TEMP.xlsx')
            rows = list(wb.worksheets[0].rows)
            wb.close()
            args['rows'] = str(len(rows))
            parsing(rows)
    return render_template("index.html", args=args)


def parsing(rows):

    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1420,1080')
    # with Xvfb() as vdisplay:
    #     vdisplay.start()
    # driver = webdriver.Chrome("chromedriver.exe", chrome_options=options)
    driver = webdriver.Chrome(chrome_options=options)

    driver.implicitly_wait(1)

    date_of_pars = datetime.utcnow()

    for row in rows:
        url = row[0].value
        if not url:
            continue

        if "/videos" not in url:
            url = '{}/videos'.format(url)

        name_of_bloger = url.split('/')[-2]
        driver.get(url)
        body = driver.find_element_by_tag_name('body')
        count = 0
        change = 0
        while not change:
            body.send_keys(Keys.END)
            time.sleep(1)
            len_prev = len(body.find_elements_by_tag_name('ytd-grid-video-renderer'))
            if count == len_prev:
                change += 1
            count += len_prev - count
        videos = []
        links = driver.find_elements_by_id('thumbnail')
        for link in links:
            try:
                href = link.get_attribute('href')
            except Exception as ex:
                error_json = json.dumps({
                    "link": link,
                })
                LOGGER.error("{}\n{}\n".format(ex, error_json))
            else:
                videos.append(href)

        rows = [
            ['Link', '#Tags', 'Name', 'Views', 'Date of publication', 'Likes', 'Dislikes', 'Date of parsing',
             'Description', 'First comment', 'Answers for first comment', 'All comments']
        ]
        element_class_name_suf = "style-scope.ytd-video-primary-info-renderer"
        for video in videos:
            row = [video, ]
            try:
                driver.get(video)
                try:
                    tags = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "super-title.{}".format(element_class_name_suf)))
                    ).text
                except Exception as ex:
                    error_json = json.dumps({
                        "video URL": video,
                    })
                    LOGGER.error("{}\n{}\n".format(ex, error_json))

                    tags = '-'

                first_text = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "{}".format(element_class_name_suf)))
                ).text.split('\n')

                if 'просмотр' not in first_text[1]:
                    first_text = first_text[1:]

                name_of_vid = first_text[0]
                element_class_name = "content.style-scope.ytd-video-secondary-info-renderer"
                description = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, element_class_name)))

                if 'просмотров' in first_text[1]:
                    count_of_watch = first_text[1].split('просмотров')[0]
                    date_of_pub = first_text[1].split('просмотров')[1]
                elif 'просмотр' in first_text[1]:
                    count_of_watch = first_text[1].split('просмотр')[0]
                    date_of_pub = first_text[1].split('просмотр')[1]
                else:
                    count_of_watch = first_text[1].split('просмотра')[0]
                    date_of_pub = first_text[1].split('просмотра')[1]
                likes = first_text[2]
                dislikes = first_text[3]

                row.extend([tags, name_of_vid, count_of_watch,
                            date_of_pub, likes, dislikes,
                            date_of_pars, description.text])

                body = driver.find_element_by_tag_name('body')

                body.send_keys(Keys.END)
                time.sleep(1)
                driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")

                count = 0
                change = 0
                while not change:
                    body.send_keys(Keys.END)
                    time.sleep(1)
                    len_com = len(body.find_elements_by_tag_name('ytd-comment-thread-renderer'))
                    if count == len_com:
                        change += 1
                    count += len_com - count

                first_comment = ''
                answers = ''
                all_comments = ''
                comments = body.find_elements_by_tag_name('ytd-comment-thread-renderer')
                if comments:
                    if 'Закреплено пользователем' in comments[0].text:
                        first_comment += comments[0].text.split('ОТВЕТИТЬ')[0]
                        more = comments[0].find_elements_by_id('replies')
                        if more:
                            try:
                                button = WebDriverWait(more[0], 2).until(EC.presence_of_element_located((By.ID, "more-replies")))
                                driver.execute_script("arguments[0].click();", button)
                                replies_block = comments[0].find_element_by_id('replies')
                                # replies = replies_block.find_elements_by_id()
                                answers += '\n\n' + replies_block.text
                            except Exception as ex:
                                error_json = json.dumps({
                                    "video URL": video,
                                })
                                LOGGER.error("{}\n{}\n".format(ex, error_json))

                    if first_comment:
                        comments = comments[1:]
                    for comment in comments:
                        all_comments += comment.text.split('ОТВЕТИТЬ')[0]
                        more = comment.find_elements_by_id('replies')
                        if more:
                            try:
                                button = WebDriverWait(more[0], 2).until(
                                    EC.presence_of_element_located((By.ID, "more-replies"))
                                )
                                driver.execute_script("arguments[0].click();", button)
                                replies_block = comment.find_element_by_id('replies')
                                all_comments += '\n\n' + replies_block.text
                            except:
                                pass
                        all_comments += '\n---------\n'
                row.append(first_comment)
                row.append(answers)
                row.append(all_comments)

                rows.append(row)
            except Exception as ex:
                error_json = json.dumps({
                    "video URL": video,
                })
                LOGGER.error("{}\n{}\n".format(ex, error_json))

        wb = openpyxl.Workbook()
        # wb = openpyxl.load_workbook(name_of_bloger + '.xlsx')
        ws = wb.active
        for row in rows:
            ws.append(row)

        wb.save(name_of_bloger + '.xlsx')


if __name__ == '__main__':
    app.run()
